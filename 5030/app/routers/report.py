from datetime import date, timedelta
from typing import Optional
import io
from urllib.parse import quote
from sqlalchemy.orm import Session
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.database import get_db
from app import models

router = APIRouter()


def _get_content_disposition(filename: str) -> dict:
    encoded_filename = quote(filename, safe="")
    ascii_filename = filename.encode("ascii", errors="replace").decode("ascii")
    return {
        "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}; filename={ascii_filename}"
    }


@router.post("/attendance/monthly/export")
def export_monthly_attendance(
    year: int,
    month: int,
    department: Optional[str] = None,
    db: Session = Depends(get_db)
):
    query = db.query(models.MonthlyAttendanceSummary).filter(
        models.MonthlyAttendanceSummary.year == year,
        models.MonthlyAttendanceSummary.month == month
    )
    
    if department:
        query = query.join(models.Employee).filter(models.Employee.department == department)
    
    summaries = query.all()
    
    if not summaries:
        raise HTTPException(status_code=404, detail="未找到该月份的考勤数据")
    
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}年{month}月考勤汇总"
    
    headers = [
        "工号", "姓名", "部门", "岗位",
        "应出勤天数", "实际出勤天数",
        "迟到天数", "早退天数", "旷工天数",
        "总工时", "是否全勤", "全勤奖"
    ]
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row_idx, summary in enumerate(summaries, 2):
        employee = db.query(models.Employee).filter(
            models.Employee.id == summary.employee_id
        ).first()
        
        ws.cell(row=row_idx, column=1, value=employee.employee_no if employee else "")
        ws.cell(row=row_idx, column=2, value=employee.name if employee else "")
        ws.cell(row=row_idx, column=3, value=employee.department if employee else "")
        ws.cell(row=row_idx, column=4, value=employee.position if employee else "")
        ws.cell(row=row_idx, column=5, value=summary.total_work_days)
        ws.cell(row=row_idx, column=6, value=summary.actual_work_days)
        ws.cell(row=row_idx, column=7, value=summary.total_late_days)
        ws.cell(row=row_idx, column=8, value=summary.total_early_leave_days)
        ws.cell(row=row_idx, column=9, value=summary.total_absent_days)
        ws.cell(row=row_idx, column=10, value=summary.total_work_hours)
        ws.cell(row=row_idx, column=11, value="是" if summary.is_full_attendance else "否")
        ws.cell(row=row_idx, column=12, value=summary.full_attendance_bonus)
        
        if summary.is_full_attendance:
            ws.cell(row=row_idx, column=11).fill = PatternFill(
                start_color="70AD47", end_color="70AD47", fill_type="solid"
            )
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    wb.save(output)
    output.seek(0)
    
    filename = f"考勤汇总_{year}年{month}月.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=_get_content_disposition(filename)
    )


@router.post("/attendance/detailed/export")
def export_detailed_attendance(
    employee_id: Optional[int] = None,
    start_date: date = None,
    end_date: date = None,
    db: Session = Depends(get_db)
):
    query = db.query(models.AttendanceRecord)
    
    if employee_id:
        query = query.filter(models.AttendanceRecord.employee_id == employee_id)
    if start_date:
        query = query.filter(models.AttendanceRecord.date >= start_date)
    if end_date:
        query = query.filter(models.AttendanceRecord.date <= end_date)
    
    records = query.order_by(models.AttendanceRecord.date).all()
    
    if not records:
        raise HTTPException(status_code=404, detail="未找到考勤记录")
    
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "考勤明细"
    
    headers = [
        "日期", "工号", "姓名", "部门",
        "上班打卡时间", "下班打卡时间",
        "工作时长", "迟到分钟", "早退分钟",
        "是否旷工", "备注"
    ]
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row_idx, record in enumerate(records, 2):
        employee = db.query(models.Employee).filter(
            models.Employee.id == record.employee_id
        ).first()
        
        ws.cell(row=row_idx, column=1, value=str(record.date))
        ws.cell(row=row_idx, column=2, value=employee.employee_no if employee else "")
        ws.cell(row=row_idx, column=3, value=employee.name if employee else "")
        ws.cell(row=row_idx, column=4, value=employee.department if employee else "")
        ws.cell(row=row_idx, column=5, value=str(record.check_in) if record.check_in else "")
        ws.cell(row=row_idx, column=6, value=str(record.check_out) if record.check_out else "")
        ws.cell(row=row_idx, column=7, value=record.work_hours)
        ws.cell(row=row_idx, column=8, value=record.late_minutes)
        ws.cell(row=row_idx, column=9, value=record.early_leave_minutes)
        ws.cell(row=row_idx, column=10, value="是" if record.is_absent else "否")
        ws.cell(row=row_idx, column=11, value=record.remarks or "")
        
        if record.late_minutes > 0:
            ws.cell(row=row_idx, column=8).fill = PatternFill(
                start_color="FFC000", end_color="FFC000", fill_type="solid"
            )
        if record.early_leave_minutes > 0:
            ws.cell(row=row_idx, column=9).fill = PatternFill(
                start_color="FFC000", end_color="FFC000", fill_type="solid"
            )
        if record.is_absent:
            ws.cell(row=row_idx, column=10).fill = PatternFill(
                start_color="FF0000", end_color="FF0000", fill_type="solid"
            )
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18
    
    wb.save(output)
    output.seek(0)
    
    filename = f"考勤明细_{date.today()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=_get_content_disposition(filename)
    )


@router.get("/salary/export")
def export_salary_data(
    year: int,
    month: int,
    db: Session = Depends(get_db)
):
    summaries = db.query(models.MonthlyAttendanceSummary).filter(
        models.MonthlyAttendanceSummary.year == year,
        models.MonthlyAttendanceSummary.month == month
    ).all()
    
    if not summaries:
        raise HTTPException(status_code=404, detail="未找到该月份的考勤数据")
    
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "薪资核算数据"
    
    headers = [
        "工号", "姓名", "部门", "岗位",
        "实际出勤天数", "迟到次数", "早退次数", "旷工天数",
        "全勤奖", "备注"
    ]
    
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row_idx, summary in enumerate(summaries, 2):
        employee = db.query(models.Employee).filter(
            models.Employee.id == summary.employee_id
        ).first()
        
        remarks = []
        if summary.total_late_days > 0:
            remarks.append(f"迟到{summary.total_late_days}次")
        if summary.total_early_leave_days > 0:
            remarks.append(f"早退{summary.total_early_leave_days}次")
        if summary.total_absent_days > 0:
            remarks.append(f"旷工{summary.total_absent_days}天")
        
        ws.cell(row=row_idx, column=1, value=employee.employee_no if employee else "")
        ws.cell(row=row_idx, column=2, value=employee.name if employee else "")
        ws.cell(row=row_idx, column=3, value=employee.department if employee else "")
        ws.cell(row=row_idx, column=4, value=employee.position if employee else "")
        ws.cell(row=row_idx, column=5, value=summary.actual_work_days)
        ws.cell(row=row_idx, column=6, value=summary.total_late_days)
        ws.cell(row=row_idx, column=7, value=summary.total_early_leave_days)
        ws.cell(row=row_idx, column=8, value=summary.total_absent_days)
        ws.cell(row=row_idx, column=9, value=summary.full_attendance_bonus)
        ws.cell(row=row_idx, column=10, value="; ".join(remarks) if remarks else "正常")
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    wb.save(output)
    output.seek(0)
    
    filename = f"薪资核算数据_{year}年{month}月.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=_get_content_disposition(filename)
    )


@router.get("/leave/export")
def export_leave_records(
    year: Optional[int] = None,
    employee_id: Optional[int] = None,
    leave_type: Optional[str] = None,
    db: Session = Depends(get_db)
):
    query = db.query(models.LeaveRequest)
    
    if year:
        query = query.filter(
            models.LeaveRequest.start_date >= f"{year}-01-01",
            models.LeaveRequest.start_date <= f"{year}-12-31"
        )
    if employee_id:
        query = query.filter(models.LeaveRequest.employee_id == employee_id)
    if leave_type:
        query = query.filter(models.LeaveRequest.leave_type == leave_type)
    
    leaves = query.order_by(models.LeaveRequest.created_at.desc()).all()
    
    if not leaves:
        raise HTTPException(status_code=404, detail="未找到请假记录")
    
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "请假记录"
    
    leave_type_names = {
        "personal": "事假",
        "sick": "病假",
        "annual": "年假",
        "compensatory": "调休",
        "marriage": "婚假",
        "maternity": "产假",
        "paternity": "陪产假",
        "bereavement": "丧假"
    }
    
    status_names = {
        "pending": "待审批",
        "approved": "已批准",
        "rejected": "已拒绝"
    }
    
    headers = [
        "工号", "姓名", "部门", "请假类型",
        "开始日期", "结束日期", "天数",
        "请假原因", "状态", "申请时间"
    ]
    
    header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row_idx, leave in enumerate(leaves, 2):
        employee = db.query(models.Employee).filter(
            models.Employee.id == leave.employee_id
        ).first()
        
        ws.cell(row=row_idx, column=1, value=employee.employee_no if employee else "")
        ws.cell(row=row_idx, column=2, value=employee.name if employee else "")
        ws.cell(row=row_idx, column=3, value=employee.department if employee else "")
        ws.cell(row=row_idx, column=4, value=leave_type_names.get(leave.leave_type, leave.leave_type))
        ws.cell(row=row_idx, column=5, value=str(leave.start_date))
        ws.cell(row=row_idx, column=6, value=str(leave.end_date))
        ws.cell(row=row_idx, column=7, value=leave.days)
        ws.cell(row=row_idx, column=8, value=leave.reason)
        ws.cell(row=row_idx, column=9, value=status_names.get(leave.status, leave.status))
        ws.cell(row=row_idx, column=10, value=str(leave.created_at))
        
        if leave.status == "approved":
            ws.cell(row=row_idx, column=9).fill = PatternFill(
                start_color="70AD47", end_color="70AD47", fill_type="solid"
            )
        elif leave.status == "rejected":
            ws.cell(row=row_idx, column=9).fill = PatternFill(
                start_color="FF0000", end_color="FF0000", fill_type="solid"
            )
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    wb.save(output)
    output.seek(0)
    
    filename = f"请假记录_{date.today()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=_get_content_disposition(filename)
    )
