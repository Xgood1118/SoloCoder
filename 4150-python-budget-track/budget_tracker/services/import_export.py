import csv
import os

from budget_tracker.database import get_db
from budget_tracker.services.category import CategoryService
from budget_tracker.services.transaction import TransactionService

try:
    import openpyxl
except ImportError:
    openpyxl = None


class ImportExportService:
    def __init__(self, db=None):
        self.db = db or get_db()
        self.category_service = CategoryService(self.db)
        self.transaction_service = TransactionService(self.db)

    def import_csv(self, file_path, account_id, delimiter=",", encoding="utf-8"):
        imported = 0
        skipped = 0
        errors = []

        with open(file_path, "r", encoding=encoding, newline="") as f:
            reader = csv.DictReader(f, delimiter=delimiter)
            for row_num, row in enumerate(reader, start=2):
                try:
                    date = row.get("date", "").strip()
                    amount_str = row.get("amount", "").strip()
                    description = row.get("description", "").strip()
                    if not date or not amount_str:
                        skipped += 1
                        continue

                    amount = float(amount_str)
                    transaction_type = "income" if amount >= 0 else "expense"
                    abs_amount = abs(amount)

                    notes = row.get("notes", "").strip()
                    tags = row.get("tags", "").strip()
                    category_name = row.get("category", "").strip()

                    category_id = None
                    if category_name:
                        cat_row = self.db.query_one("SELECT id FROM categories WHERE name = ?", (category_name,))
                        if cat_row:
                            category_id = cat_row["id"]
                    if category_id is None:
                        category_id = self.category_service.auto_classify(description)

                    self.transaction_service.create_transaction(
                        account_id=account_id,
                        transaction_type=transaction_type,
                        amount=abs_amount,
                        category_id=category_id,
                        description=description,
                        notes=notes,
                        tags=tags,
                        date=date,
                    )
                    imported += 1
                except Exception as e:
                    errors.append({"row": row_num, "error": str(e)})

        return {"imported": imported, "skipped": skipped, "errors": errors}

    def export_csv(self, account_id=None, start_date=None, end_date=None, file_path="export.csv"):
        transactions = self.transaction_service.list_transactions(
            account_id=account_id,
            start_date=start_date,
            end_date=end_date,
            limit=None,
            offset=0,
        )

        with open(file_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["id", "date", "transaction_type", "amount", "category", "description", "notes", "tags", "account"])
            for txn in transactions:
                category_name = ""
                if txn.category_id:
                    cat = self.category_service.get_category(txn.category_id)
                    if cat:
                        category_name = cat.name
                account_name = ""
                acc_row = self.db.query_one("SELECT name FROM accounts WHERE id = ?", (txn.account_id,))
                if acc_row:
                    account_name = acc_row["name"]
                writer.writerow([
                    txn.id,
                    txn.date,
                    txn.transaction_type,
                    txn.amount,
                    category_name,
                    txn.description,
                    txn.notes,
                    txn.tags,
                    account_name,
                ])

        return file_path

    def export_excel(self, account_id=None, start_date=None, end_date=None, file_path="export.xlsx"):
        if openpyxl is None:
            raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")

        transactions = self.transaction_service.list_transactions(
            account_id=account_id,
            start_date=start_date,
            end_date=end_date,
            limit=None,
            offset=0,
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"

        headers = ["id", "date", "transaction_type", "amount", "category", "description", "notes", "tags", "account"]
        bold_font = openpyxl.styles.Font(bold=True)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = bold_font

        for row_idx, txn in enumerate(transactions, start=2):
            category_name = ""
            if txn.category_id:
                cat = self.category_service.get_category(txn.category_id)
                if cat:
                    category_name = cat.name
            account_name = ""
            acc_row = self.db.query_one("SELECT name FROM accounts WHERE id = ?", (txn.account_id,))
            if acc_row:
                account_name = acc_row["name"]

            ws.cell(row=row_idx, column=1, value=txn.id)
            ws.cell(row=row_idx, column=2, value=txn.date)
            ws.cell(row=row_idx, column=3, value=txn.transaction_type)
            amount_cell = ws.cell(row=row_idx, column=4, value=txn.amount)
            amount_cell.number_format = "#,##0.00"
            ws.cell(row=row_idx, column=5, value=category_name)
            ws.cell(row=row_idx, column=6, value=txn.description)
            ws.cell(row=row_idx, column=7, value=txn.notes)
            ws.cell(row=row_idx, column=8, value=txn.tags)
            ws.cell(row=row_idx, column=9, value=account_name)

        for col_idx, header in enumerate(headers, start=1):
            max_length = len(str(header))
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
                for value in row:
                    if value is not None:
                        max_length = max(max_length, len(str(value)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)

        wb.save(file_path)
        return file_path
