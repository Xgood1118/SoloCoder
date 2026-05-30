from __future__ import annotations

from datetime import datetime, date
from typing import Any, Dict, List, Optional

from data_io.writer.base import BaseWriter
from data_io.utils.helpers import ensure_dir


class ExcelWriter(BaseWriter):
    def __init__(
        self,
        file_path: str,
        headers: Optional[List[str]] = None,
        encoding: str = "utf-8",
        sheet_name: str = "Sheet1",
        append: bool = False,
        auto_filter: bool = False,
        freeze_header: bool = False,
        **kwargs,
    ):
        super().__init__(file_path, headers, encoding, **kwargs)
        self.sheet_name = sheet_name
        self.append = append
        self.auto_filter = auto_filter
        self.freeze_header = freeze_header
        self._wb = None
        self._ws = None
        self._row_count = 0
        self._header_written = False

    def _init_writer(self) -> None:
        ensure_dir(self.file_path)
        from openpyxl import Workbook, load_workbook
        self._header_written = False
        self._row_count = 0
        if self.append:
            try:
                self._wb = load_workbook(self.file_path)
                if self.sheet_name in self._wb.sheetnames:
                    self._ws = self._wb[self.sheet_name]
                    self._row_count = self._ws.max_row
                    self._header_written = self._row_count > 0
                else:
                    self._ws = self._wb.create_sheet(self.sheet_name)
            except (FileNotFoundError, OSError):
                self._wb = Workbook(write_only=True)
                self._ws = self._wb.create_sheet(self.sheet_name)
        else:
            self._wb = Workbook(write_only=True)
            self._ws = self._wb.create_sheet(self.sheet_name)

    def _write_header(self) -> None:
        if not self.headers:
            return
        if not self._header_written:
            self._ws.append(self.headers)
            self._row_count += 1
            self._header_written = True
        if self.freeze_header:
            try:
                self._ws.freeze_panes = "A2"
            except AttributeError:
                pass
        if self.auto_filter:
            pass

    def _convert_value(self, val: Any) -> Any:
        if val is None:
            return None
        if isinstance(val, (int, float, bool, datetime, date)):
            return val
        return str(val)

    def _write_row(self, data: Dict[str, Any]) -> None:
        if not self.headers:
            self.headers = list(data.keys())
            self._write_header()
        row = []
        for h in self.headers:
            val = data.get(h)
            row.append(self._convert_value(val))
        self._ws.append(row)
        self._row_count += 1

    def _finalize(self) -> None:
        if self._wb:
            if self.auto_filter and self._row_count > 1 and not self._wb.write_only:
                try:
                    last_col = chr(ord('A') + len(self.headers) - 1) if self.headers else 'Z'
                    self._ws.auto_filter.ref = f"A1:{last_col}{self._row_count}"
                except Exception:
                    pass
            self._wb.save(self.file_path)
            self._wb = None
            self._ws = None

    def set_column_width(self, column: str, width: int) -> None:
        if self._ws:
            self._ws.column_dimensions[column].width = width

    def set_cell_style(self, row: int, column: str, **kwargs) -> None:
        if self._ws:
            from openpyxl.styles import Font, PatternFill, Alignment, Border
            cell = self._ws[f"{column}{row}"]
            if "font" in kwargs:
                cell.font = kwargs["font"]
            if "fill" in kwargs:
                cell.fill = kwargs["fill"]
            if "alignment" in kwargs:
                cell.alignment = kwargs["alignment"]
