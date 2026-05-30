from __future__ import annotations

from datetime import datetime, date
from typing import Any, Dict, Iterator, List, Optional

from data_io.reader.base import BaseReader
from data_io.models import Record


def _convert_excel_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if value == int(value):
            return int(value)
    return value


class ExcelReader(BaseReader):
    def __init__(
        self,
        file_path: str,
        encoding: Optional[str] = None,
        sheet_name: Optional[str] = None,
        **kwargs,
    ):
        super().__init__(file_path, encoding, **kwargs)
        self.sheet_name = sheet_name
        self._headers: Optional[List[str]] = None
        self._total: Optional[int] = None

    def _get_workbook(self):
        from openpyxl import load_workbook

        return load_workbook(self.file_path, read_only=True, data_only=True)

    def _get_sheet(self, wb):
        if self.sheet_name:
            return wb[self.sheet_name]
        return wb.active

    def _resolve_merged_cells(self, ws) -> Dict:
        merged = {}
        if not hasattr(ws, "merged_cells"):
            return merged
        for merge_range in ws.merged_cells.ranges:
            min_row = merge_range.min_row
            min_col = merge_range.min_col
            try:
                cell_value = ws.cell(row=min_row, column=min_col).value
            except Exception:
                cell_value = None
            for row in range(merge_range.min_row, merge_range.max_row + 1):
                for col in range(merge_range.min_col, merge_range.max_col + 1):
                    if row != min_row or col != min_col:
                        merged[(row, col)] = cell_value
        return merged

    def read_headers(self) -> List[str]:
        if self._headers is not None:
            return self._headers
        wb = self._get_workbook()
        try:
            ws = self._get_sheet(wb)
            self._headers = []
            for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                val = _convert_excel_value(cell.value) if cell.value else str(cell.column)
                self._headers.append(str(val).strip())
        finally:
            wb.close()
        return self._headers

    def read_rows(
        self,
        start: int = 0,
        batch_size: Optional[int] = None,
    ) -> Iterator[Record]:
        headers = self.read_headers()
        wb = self._get_workbook()
        try:
            ws = self._get_sheet(wb)
            merged = self._resolve_merged_cells(ws)
            row_index = 0
            yielded = 0
            for row in ws.iter_rows(min_row=2):
                data = {}
                all_empty = True
                for i, cell in enumerate(row):
                    if i >= len(headers):
                        break
                    val = merged.get((cell.row, cell.column), cell.value)
                    val = _convert_excel_value(val)
                    if val is not None and val != "":
                        all_empty = False
                    data[headers[i]] = val
                if self._kwargs.get("skip_empty_rows", True) and all_empty:
                    row_index += 1
                    continue
                if row_index < start:
                    row_index += 1
                    continue
                yield Record(row_index=row_index, data=data)
                row_index += 1
                yielded += 1
                if batch_size is not None and yielded >= batch_size:
                    break
        finally:
            wb.close()

    def total_rows(self) -> int:
        if self._total is not None:
            return self._total
        wb = self._get_workbook()
        try:
            ws = self._get_sheet(wb)
            self._total = ws.max_row - 1 if ws.max_row > 0 else 0
        finally:
            wb.close()
        return self._total


class XlsReader(BaseReader):
    def __init__(
        self,
        file_path: str,
        encoding: Optional[str] = None,
        sheet_name: Optional[str] = None,
        **kwargs,
    ):
        super().__init__(file_path, encoding, **kwargs)
        self.sheet_name = sheet_name
        self._headers: Optional[List[str]] = None
        self._total: Optional[int] = None

    def _get_workbook(self):
        import xlrd

        return xlrd.open_workbook(self.file_path)

    def _get_sheet(self, wb):
        if self.sheet_name:
            return wb.sheet_by_name(self.sheet_name)
        return wb.sheet_by_index(0)

    def _convert_xls_value(self, wb, cell) -> Any:
        if cell.ctype == xlrd.XL_CELL_TEXT:
            return cell.value.strip()
        elif cell.ctype == xlrd.XL_CELL_NUMBER:
            val = cell.value
            if val == int(val):
                return int(val)
            return val
        elif cell.ctype == xlrd.XL_CELL_DATE:
            dt = xlrd.xldate_as_tuple(cell.value, wb.datemode)
            from datetime import datetime
            return datetime(*dt).strftime("%Y-%m-%d %H:%M:%S")
        elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
            return bool(cell.value)
        elif cell.ctype == xlrd.XL_CELL_EMPTY:
            return None
        return cell.value

    def read_headers(self) -> List[str]:
        if self._headers is not None:
            return self._headers
        import xlrd
        wb = self._get_workbook()
        ws = self._get_sheet(wb)
        self._headers = []
        for col in range(ws.ncols):
            cell = ws.cell(0, col)
            val = self._convert_xls_value(wb, cell)
            self._headers.append(str(val).strip() if val is not None else str(col))
        return self._headers

    def read_rows(
        self,
        start: int = 0,
        batch_size: Optional[int] = None,
    ) -> Iterator[Record]:
        import xlrd
        headers = self.read_headers()
        wb = self._get_workbook()
        ws = self._get_sheet(wb)
        row_index = 0
        yielded = 0
        for r in range(1, ws.nrows):
            data = {}
            all_empty = True
            for c in range(ws.ncols):
                if c >= len(headers):
                    break
                val = self._convert_xls_value(wb, ws.cell(r, c))
                if val is not None and val != "":
                    all_empty = False
                data[headers[c]] = val
            if self._kwargs.get("skip_empty_rows", True) and all_empty:
                row_index += 1
                continue
            if row_index < start:
                row_index += 1
                continue
            yield Record(row_index=row_index, data=data)
            row_index += 1
            yielded += 1
            if batch_size is not None and yielded >= batch_size:
                break

    def total_rows(self) -> int:
        if self._total is not None:
            return self._total
        wb = self._get_workbook()
        ws = self._get_sheet(wb)
        self._total = max(0, ws.nrows - 1)
        return self._total
